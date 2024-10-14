import requests
from requests.auth import HTTPBasicAuth
import pandas as pd
from jinja2 import Template
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill
import logging
import sys

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger()

class JiraExcelFormatter:
    def __init__(self, jira_url, email, api_token, output_file):
        self.jira_url = jira_url
        self.email = email
        self.api_token = api_token
        self.output_file = output_file

    def fetch_jira_data(self):
        """Fetch data from Jira and return task list"""
        logger.info("Fetching data from Jira...")
        try:
            # Define the list of all field IDs as per your requirements
            field_ids = [
                'statuscategorychangedate', 'parent', 'fixVersions', 'statusCategory', 'resolution', 
                'customfield_10112', 'customfield_10113', 'customfield_10114', 'customfield_10104', 
                'lastViewed', 'priority', 'labels', 'customfield_10103', 'timeestimate', 
                'aggregatetimeoriginalestimate', 'versions', 'issuelinks', 'assignee', 'status', 
                'components', 'issuekey', 'aggregatetimeestimate', 'creator', 'subtasks', 
                'reporter', 'aggregateprogress', 'progress', 'votes', 'worklog', 'issuetype', 
                'timespent', 'customfield_10030', 'project', 'customfield_10031', 'customfield_10032', 
                'customfield_10033', 'aggregatetimespent', 'customfield_10034', 'customfield_10035', 
                'customfield_10037', 'customfield_10027', 'customfield_10028', 'customfield_10029', 
                'resolutiondate', 'workratio', 'watches', 'issuerestriction', 'thumbnail', 
                'created', 'customfield_10020', 'customfield_10021', 'customfield_10022', 
                'customfield_10023', 'customfield_10024', 'customfield_10025', 'customfield_10026', 
                'customfield_10016', 'customfield_10017', 'customfield_10018', 'customfield_10019', 
                'updated', 'timeoriginalestimate', 'description', 'customfield_10010', 
                'customfield_10011', 'customfield_10012', 'customfield_10013', 'customfield_10014', 
                'timetracking', 'customfield_10015', 'customfield_10005', 'customfield_10126', 
                'customfield_10006', 'security', 'customfield_10007', 'customfield_10008', 
                'attachment', 'customfield_10009', 'summary', 'customfield_10120', 
                'customfield_10000', 'customfield_10121', 'customfield_10122', 'customfield_10001', 
                'customfield_10123', 'customfield_10002', 'customfield_10003', 'customfield_10124', 
                'customfield_10125', 'customfield_10004', 'customfield_10115', 'customfield_10116', 
                'environment', 'customfield_10117', 'customfield_10118', 'customfield_10119', 
                'duedate', 'comment'
            ]

            # Convert the list of field IDs into a comma-separated string
            fields = ','.join(field_ids)

            # Define the query for the DevOps project, including the custom fields
            jql_query = {
                'jql': 'project="DevOps"',  # Modify project key as needed
                'fields': field_ids
            }

            # Make the API request
            response = requests.post(
                self.jira_url,
                json=jql_query,
                auth=HTTPBasicAuth(self.email, self.api_token),
                headers={'Content-Type': 'application/json'}
            )

            # Check for a successful response
            if response.status_code == 200:
                logger.info("Data successfully retrieved from Jira")
                tasks = response.json()
                return tasks['issues']
            else:
                logger.error(f"Failed to retrieve tasks, status code: {response.status_code}")
                logger.error(response.text)
                sys.exit(1)

        except Exception as e:
            logger.exception("An error occurred while fetching data from Jira")
            sys.exit(1)

    def prepare_task_list(self, issues):
        """Prepare task list for Excel"""
        logger.info("Preparing task list for Excel...")
        task_list = []
        for issue in issues:
            task_list.append({
                'Task Key': issue['key'],
                'Summary': issue['fields'].get('summary', ''),
                'Status': issue['fields']['status']['name'],
                'Category': issue['fields'].get('customfield_10035', {}).get('value', 'No Category'),  # Category field
                'Assignee': issue['fields']['assignee']['displayName'] if issue['fields'].get('assignee') else 'Unassigned',
                'Due Date': issue['fields'].get('duedate', ''),
                'Priority': issue['fields']['priority']['name'] if 'priority' in issue['fields'] else 'No Priority',
                'Labels': ', '.join(issue['fields'].get('labels', [])) if issue['fields'].get('labels') else 'No Labels',
                'Created': issue['fields'].get('created', ''),
                'Updated': issue['fields'].get('updated', ''),
                'Reporter': issue['fields']['reporter']['displayName'] if issue['fields'].get('reporter') else 'No Reporter',
                'Team': issue['fields'].get('customfield_10001', {}).get('name', 'No Team'),  # Custom field for Team
    
                # Newly added fields
                'Status Category Changed': issue['fields'].get('statuscategorychangedate', 'No Status Category Changed'),
                'Parent': issue['fields'].get('parent', {}).get('key', 'No Parent'),
                'Fix Versions': ', '.join([ver['name'] for ver in issue['fields'].get('fixVersions', [])]) if issue['fields'].get('fixVersions') else 'No Fix Versions',
                
                # Check if 'resolution' exists and is not None
                'Resolution': issue['fields'].get('resolution', {}).get('name', 'Unresolved') if issue['fields'].get('resolution') else 'Unresolved',
    
                'DEV Completion': issue['fields'].get('customfield_10112', 'No DEV Completion'),
                'QA Completion': issue['fields'].get('customfield_10113', 'No QA Completion'),
                'Demo Deployment': issue['fields'].get('customfield_10114', 'No Demo Deployment'),
                'Remaining Estimate': issue['fields'].get('timeestimate', 'No Remaining Estimate'),
                'Σ Original Estimate': issue['fields'].get('aggregatetimeoriginalestimate', 'No Σ Original Estimate'),
                'Affects Versions': ', '.join([ver['name'] for ver in issue['fields'].get('versions', [])]) if issue['fields'].get('versions') else 'No Affects Versions',
                'Linked Issues': ', '.join([link['outwardIssue']['key'] for link in issue['fields'].get('issuelinks', []) if 'outwardIssue' in link]) if issue['fields'].get('issuelinks') else 'No Linked Issues',
                'Creator': issue['fields'].get('creator', {}).get('displayName', 'No Creator'),
                'Sub-tasks': ', '.join([sub['key'] for sub in issue['fields'].get('subtasks', [])]) if issue['fields'].get('subtasks') else 'No Sub-tasks',
                'Progress': issue['fields'].get('progress', {}).get('progress', 'No Progress'),
                'Votes': issue['fields'].get('votes', {}).get('votes', 'No Votes'),
                'Log Work': ', '.join([log['timeSpent'] for log in issue['fields'].get('worklog', {}).get('worklogs', [])]) if issue['fields'].get('worklog') else 'No Work Log',
                'Time Spent': issue['fields'].get('timespent', 'No Time Spent'),
                'Resolved': issue['fields'].get('resolutiondate', 'Not Resolved'),
                'Work Ratio': issue['fields'].get('workratio', 'No Work Ratio'),
                'Watchers': issue['fields'].get('watches', {}).get('watchCount', 'No Watchers'),
                'Images': issue['fields'].get('thumbnail', {}).get('name', 'No Images'),
                'Sprint': ', '.join([sprint['name'] for sprint in issue['fields'].get('customfield_10020', [])]) if issue['fields'].get('customfield_10020') else 'No Sprint',
                'Flagged': 'Yes' if issue['fields'].get('customfield_10021') else 'No Flag',
                'Original Estimate': issue['fields'].get('timeoriginalestimate', 'No Original Estimate'),
                'Description': issue['fields'].get('description', 'No Description'),
                'Epic Link': issue['fields'].get('customfield_10014', 'No Epic Link'),
                'Time Tracking': issue['fields'].get('timetracking', {}).get('originalEstimate', 'No Time Tracking'),
                'Environment': issue['fields'].get('environment', 'No Environment'),
                'Due date': issue['fields'].get('duedate', 'No Due date'),
            })
        return task_list

    def save_to_excel(self, task_list):
        """Save task list to an Excel file with formatting"""
        try:
            # Prepare data for Excel
            logger.info("Preparing data for Excel...")
            df = pd.DataFrame(task_list)

            # Save DataFrame to Excel
            logger.info(f"Saving data to {self.output_file}...")
            df.to_excel(self.output_file, index=False)

        except Exception as e:
            logger.exception("An error occurred while saving data to Excel")
            sys.exit(1)
            
    def format_excel(self):
        """Format the Excel file with priority colors, alignment, borders, filters, and gray header row"""
        try:
            logger.info("Formatting Excel file...")
    
            # Load the workbook to format it with openpyxl
            wb = load_workbook(self.output_file)
            ws = wb.active
    
            # Define border style
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
            # Define color fills for different priorities
            priority_colors = {
                'Medium': PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid'),  # Orange for Medium
                'High': PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),    # Red for High
                'Critical': PatternFill(start_color='8B0000', end_color='8B0000', fill_type='solid'),  # Dark Red for Critical
                'Low': PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')      # Blue for Low
            }
    
            # Define gray color for header row
            header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')  # Gray color
    
            # Apply center alignment, borders, and conditional formatting for priorities
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):  # Iterate over rows, starting from the second row to skip header
                priority_cell = row[6]  # Assuming column G is the priority column
                priority_value = str(priority_cell.value).strip()  # Ensure that we are stripping whitespace or extra characters
    
                # Print for debugging: Check what values are being read
                logger.debug(f"Row {priority_cell.row}: Priority value = '{priority_value}'")
    
                # Center align cells and apply borders
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')  # Center align cells
                    cell.border = thin_border  # Apply border to each cell
    
                # Apply color fill only in the "Priority" column (column G)
                if priority_value in priority_colors:
                    logger.info(f"Applying color for Priority '{priority_value}' at row {priority_cell.row}")
                    priority_cell.fill = priority_colors[priority_value]  # Apply color based on priority
    
            # Apply gray color and border to the header row (first row)
            logger.info("Applying gray color to the header row...")
            for cell in ws[1]:  # Targeting the first row
                cell.fill = header_fill  # Apply gray fill
                cell.alignment = Alignment(horizontal='center', vertical='center')  # Center align header cells
                cell.border = thin_border  # Apply borders to header cells
    
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter  # Get the column name
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)  # Determine the max length in the column
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width  # Set column width to fit contents
    
            # Apply filters to the header row (first row)
            logger.info("Applying filters to header row...")
            ws.auto_filter.ref = ws.dimensions  # Apply filter to the full range of data
    
            # Save the updated workbook
            wb.save(self.output_file)
            logger.info(f"Excel file saved and formatted at {self.output_file}")
    
        except Exception as e:
            logger.exception("An error occurred while formatting the Excel file")
            sys.exit(1)

class JiraUpdaterFromExcel:
    def __init__(self, jira_url, email, api_token, excel_file):
        self.jira_url = jira_url
        self.email = email
        self.api_token = api_token
        self.excel_file = excel_file

    def read_excel(self):
        """Read the Excel file and return the data as a DataFrame."""
        try:
            logger.info(f"Reading Excel file: {self.excel_file}")
            df = pd.read_excel(self.excel_file)
            return df
        except Exception as e:
            logger.exception(f"An error occurred while reading the Excel file: {self.excel_file}")
            sys.exit(1)

    def get_assignee_account_id(self, display_name):
        """Fetch the Jira account ID for the assignee by display name."""
        url = f"{self.jira_url}/rest/api/3/user/search?query={display_name}"
        try:
            response = requests.get(
                url,
                auth=HTTPBasicAuth(self.email, self.api_token),
                headers={'Content-Type': 'application/json'}
            )

            if response.status_code == 200:
                users = response.json()
                if users:
                    # Return the first match (you can refine the search logic if needed)
                    return users[0]['accountId']
                else:
                    logger.error(f"Assignee '{display_name}' not found in Jira.")
                    return None
            else:
                logger.error(f"Failed to retrieve user account ID for {display_name}, status code: {response.status_code}")
                logger.error(response.text)
                return None
        except Exception as e:
            logger.exception(f"An error occurred while fetching account ID for {display_name}")
            return None

    def update_jira_issue(self, issue_key, update_data):
        """Update the Jira issue using the Jira REST API."""
        url = f"{self.jira_url}/rest/api/2/issue/{issue_key}"

        try:
            headers = {
                'Content-Type': 'application/json'
            }

            # Make the API request to update the Jira issue
            response = requests.put(
                url,
                json=update_data,
                auth=HTTPBasicAuth(self.email, self.api_token),
                headers=headers
            )

            if response.status_code == 204:
                logger.info(f"Issue {issue_key} updated successfully.")
            else:
                logger.error(f"Failed to update issue {issue_key}, status code: {response.status_code}")
                logger.error(response.text)

        except Exception as e:
            logger.exception(f"An error occurred while updating Jira issue {issue_key}")

    def sanitize_data(self, value):
        """Helper function to sanitize data before updating Jira."""
        return str(value).strip() if value else None

    def generate_description(self, summary, category, team):
        """Generate the Jira issue description as a markdown table."""
        template = Template('''
        h2. Task Details

        || Item || Description ||
        | *Summary* | {{ summary }} |
        | *Category* | {{ category }} |
        | *Team* | {{ team }} |
        ''')
        return template.render(summary=summary, category=category, team=team)

    def process_and_update_issues(self):
        """Process the Excel file and update the corresponding Jira issues."""
        # Read the Excel file
        df = self.read_excel()

        # Iterate over the DataFrame rows and update the Jira issues
        for index, row in df.iterrows():
            issue_key = row.get('Task Key', None)
            priority = row.get('Priority', None)
            summary = row.get('Summary', None)
            assignee = row.get('Assignee', None)
            category = row.get('Category', None)  # Assuming this is the column for Category
            team = row.get('Team', None)  # Assuming this is the column for Team

            # Skip if mandatory fields are missing
            if not issue_key or not priority or not summary or not assignee or not category or not team:
                logger.error(f"Missing required fields for issue {issue_key}. Skipping this issue.")
                continue

            # Get the account ID of the assignee
            assignee_account_id = self.get_assignee_account_id(assignee)

            if not assignee_account_id:
                logger.error(f"Assignee '{assignee}' not found, skipping update for issue {issue_key}.")
                continue

            # Generate the formatted description with only Summary, Category, and Team
            description = self.generate_description(summary, category, team)

            # Prepare the data to update in Jira
            update_data = {
                "fields": {
                    "summary": self.sanitize_data(summary),
                    "priority": {"name": self.sanitize_data(priority)},
                    "assignee": {"accountId": assignee_account_id},  # Use account ID for assignment
                    "description": description  # Adding the formatted description with specific fields
                }
            }

            logger.info(f"Updating issue {issue_key} with data: {update_data}")

            # Update the Jira issue
            self.update_jira_issue(issue_key, update_data)

def main():
    try:              
        output_file = 'jira_tasks.xlsx'
        base_url = "https://myworkspace.atlassian.net"  # Correct base URL
        jira_search_url = f"{base_url}/rest/api/2/search"  # URL for fetching data
        email = "ysername@example.com"
        api_token = "xxxxxxxxxxxxxxxxxxxxx"

        user_choice = input("Choose an action: Type 'fetch' to fetch Jira data or 'update' to update Jira issues from the Excel file: ").strip().lower()

        if user_choice == 'fetch':
            # Create an instance of the JiraExcelFormatter class with the search URL
            jira_formatter = JiraExcelFormatter(jira_search_url, email, api_token, output_file)
    
            # Fetch Jira data
            issues = jira_formatter.fetch_jira_data()
    
            # Prepare task list
            task_list = jira_formatter.prepare_task_list(issues)
    
            # Save data to Excel
            jira_formatter.save_to_excel(task_list)
    
            # Format the Excel file
            jira_formatter.format_excel()            
            
        elif user_choice == 'update':
            # Use base URL for updating issues
            jira_updater = JiraUpdaterFromExcel(base_url, email, api_token, output_file)
            jira_updater.process_and_update_issues()

        else:
            logger.error("Invalid choice. Please type 'fetch' or 'update'.")
            sys.exit(1)

    except Exception as e:
        logger.exception("An unexpected error occurred in the main function")
        sys.exit(1)

if __name__ == "__main__":
    main()
